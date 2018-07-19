import glob
import os
import time
import openpyxl
from helper import common_os_helper



class Inspect_data:
    Players = []

    def find_file_name(self, name_string, place):
        # name_string 폴더의 이름을 string형태로 저장
        # place home이라면 H, away라면 A를 저장

        # serial 번호와 등번호가 적힌 .csv file의 이름을 찾음
        name_list = name_string.split('-')
        value_list = name_list[1].split('_')

        new_name = name_list[0] + '-' + str(int(value_list[0]))
        if value_list[1] == 'U17':
            new_name = new_name + '_U17'
        new_name = new_name+'_'+place+'_'+value_list[2]
        new_name = new_name.replace('/', '.xlsx')
        return new_name

    def time_calculate(self, time_string, term_min_int, type_list):
        # time_string 기존 시간을 "시간:분" 형식으로 저장
        # term_min_int 분 단위의 증가하고 싶은 시간
        # type_list return type을 list형으로 하려면 true, string형으로 하려면 false

        # time_string으로부터 term_min_int만큼 지난 시간을 return
        time_list = time_string.split(':')
        time_list[1] = int(time_list[1]) + term_min_int
        time_list[0] = int(time_list[0])
        while int(time_list[1]) >= 60:
            time_list[0] = int(time_list[0]) + 1
            time_list[1] = int(time_list[1]) - 60
        if type_list: return time_list
        else:
            time_str = str(time_list[0])+':'+str(time_list[1])
            return time_str

    def time_comparision(self, time_log, time_xl, term):
        # target_list 비교하고자 하는 목적의 시간을 string으로 저장
        # comparision_target_smaller 비교할 범위 중 작은 시간을 string으로 저장
        # comparision_target_larger 비교할 범위 중 큰 시간을 string으로 저장

        # term이 -1인 경우는 본 에러를 확인하지 않겠다는 뜻이므로 다음에러로 넘어가는 1을 return
        if term == -1:
            return 1

        list_value = self.time_calculate(time_log, term, True)
        comparision_value_smaller = self.time_calculate(time_xl, 0, True)
        comparision_value_larger = self.time_calculate(time_xl, term * 2, True)

        # 오전과 오후의 시간이 다르게 적히더라도 같게 인식하고자 함
        if list_value[0]-comparision_value_larger[0] > 9 and list_value[0]-comparision_value_smaller[0] > 9:
            comparision_value_smaller[0] += 12
            comparision_value_larger[0] += 12

        # 비교할 범위보다 작은 쪽으로 벗어나면 -1, 큰쪽으로 벗어나면 1, 범위에 포함되면 0을 return
        if comparision_value_smaller[0]>list_value[0]:
            return -1
        if comparision_value_larger[0]<list_value[0]:
            return 1
        if comparision_value_smaller[0]==list_value[0] and comparision_value_smaller[1]>list_value[1]:
            return -1
        if comparision_value_larger[0]==list_value[0] and comparision_value_larger[1]<list_value[1]:
            return 1
        return 0

    def list_compare(self, info_list, log_list, error_list, error_code, ck_time = True):
        # info_list, log_list 비교할 두 개의 list
        # error_list 기존의 에러를 담고있는 list
        # error_code 에러가 발생할 경우 각 error case에 따른 설명이 저장된 list
        # ck_time time에 대해 비교할 것이라면 true, 아니면 false

        # 각 list를 정렬 하여 한번만에 비교
        info_list.sort()
        log_list.sort()

        info_id = 0
        log_id = 0

        while info_id < len(info_list) and log_id < len(log_list):
            if int(info_list[info_id][0]) < int(log_list[log_id][0]):
                error_list.append(error_code[0][0]+str(info_list[info_id][0])+error_code[0][1])
                info_id += 1
                continue
            if int(info_list[info_id][0]) > int(log_list[log_id][0]):
                error_list.append(error_code[1][0]+str(log_list[log_id][0])+error_code[1][1])
                log_id += 1
                continue
            if ck_time and self.time_comparision(log_list[log_id][1], info_list[info_id][1], 3) != 0:
                error_list.append(error_code[2][0]+str(info_list[info_id][0])+error_code[2][1])
            elif not ck_time:
                self.Players.append([info_list[info_id][0], '', ''])
            log_id += 1
            info_id += 1

        # 아직 index를 모두 둘러보지 못한 경우 모두 error
        while info_id < len(info_list):
            error_list.append(error_code[0][0]+str(info_list[info_id][0])+error_code[0][1])
            info_id += 1
        while log_id < len(log_list):
            error_list.append(error_code[1][0]+str(log_list[log_id][0])+error_code[1][1])
            log_id += 1

        return error_list

    def set_players_runtime(self, in_list, out_list):
        # in_list sub하였을 때 들어온 선수의 번호와 시간이 저장된 list
        # out_list sub 하였을 때 나간 선수의 번호와 시간이 저장된 list

        # self.Players list에 [등번호, 들어온 시간, 나간 시간]을 각 player 별로 append

        # .index를 이용해 search를 진행하기 위해 시간과 등번호 분리
        in_player = []
        in_time = []
        out_player = []
        out_time = []
        for in_val in in_list:
            in_player.append(in_val[0])
            in_time.append(in_val[1])
        for out_val in out_list:
            out_player.append(out_val[0])
            out_time.append(out_val[1])

        for i in range(len(self.Players)):
            try:
                num = in_player.index(self.Players[i][0])
                self.Players[i][1] = in_time[num]
            except ValueError:
                self.Players[i][1] = self.START1
            try:
                num = out_player.index(self.Players[i][0])
                self.Players[i][2] = out_time[num]
            except ValueError:
                self.Players[i][2] = self.END2

    def write_error(self, path_write_folder, file_name, write_order, error_list):
        # error를 file에 기록
        common_os_helper.create_dir(path_write_folder)
        file_to_write = open(path_write_folder + file_name, 'w', encoding="utf-8")
        file_to_write.write(write_order)
        for error_str in error_list:
            file_to_write.write(error_str + '\n')
        file_to_write.close()

    def find_diff_log(self, path_read_xl_folder, path_read_log_folder, path_write_folder, name_of_dir):
        # excel file과 log.csv의 다른 부분을 error로 판단하여 file에 기록

        path_read_xl_folder = common_os_helper.check_slash(path_read_xl_folder)
        path_read_log_folder = common_os_helper.check_slash(path_read_log_folder)
        path_write_folder = common_os_helper.check_slash(path_write_folder)
        name_of_dir = common_os_helper.check_slash(name_of_dir)
        path_write_folder = path_write_folder+name_of_dir

        file_to_read_log = open(path_read_log_folder+name_of_dir+'log.csv')
        log_info_lists = file_to_read_log.readlines()

        # 경기가 home인지, away인지 모르기 때문에 둘 다 시행
        try:
            file_to_read_xl = openpyxl.load_workbook(path_read_xl_folder+self.find_file_name(name_of_dir,'H'))
        except:
            file_to_read_xl = openpyxl.load_workbook(path_read_xl_folder+self.find_file_name(name_of_dir, 'A'))

        active_sheet = file_to_read_xl.active
        player_sub = active_sheet['A11':'C17']
        start_time1 = active_sheet['A7'].value
        end_time1 = active_sheet['B7'].value
        start_time2 = active_sheet['D7'].value
        end_time2 = active_sheet['E7'].value

        start1 = log_info_lists[1].split(',')[1]
        self.START1 = start1
        error = self.time_comparision(start1, start_time1, 3)

        error_list = []
        if error != 0:
            error_list.append('TimeDiff,START1,incorrect')

        sub_in_list = []
        sub_out_list = []
        for row in player_sub:
            if row[0].value != row[1].value:
                time_list = row[0].value.split('+')

                time_int = 0
                for time_val in time_list:
                    time_int += int(time_val)

                if int(time_list[0]) <= 45:
                    time_str = self.time_calculate(start_time1, time_int, False)
                else:
                    time_str = self.time_calculate(start_time2, time_int-45, False)
                sub_in_list.append([int(row[1].value), time_str])
                sub_out_list.append([int(row[2].value), time_str])

        # log.csv에서 읽어온 내용과 excel file의 내용 비교
        sub_log_in_list = []
        sub_log_out_list = []
        for log_info in log_info_lists[2:]:
            log_list = log_info.split(',')
            if log_list[0] == 'END1':
                self.END1 = log_list[1]
                if self.time_comparision(log_list[1], end_time1, 3) != 0:
                    error_list.append('TimeDiff,END1,incorrect')
            if log_list[0] == 'START2':
                self.START2 = log_list[1]
                if self.time_comparision(log_list[1], start_time2, 3) != 0:
                    error_list.append('TimeDiff,START2,incorrect')
            if log_list[0] == 'END2':
                self.END2 = log_list[1]
                if self.time_comparision(log_list[1], end_time2, 3) != 0:
                    error_list.append('TimeDiff,END2,incorrect')
            if log_list[0] == 'SUB':
                try:
                    sub_log_in_list.append([int(log_list[3]), log_list[1]])
                    sub_log_out_list.append([int(log_list[2]), log_list[1]])
                except:
                    try:
                        sub_log_out_list.append([int(log_list[2]), log_list[1]])
                    except:
                        continue

        self.set_players_runtime(sub_log_in_list, sub_log_out_list)

        # sub된 선수들에 대해 log.csv와 excel file이 다른 부분이 있는지 확인, 기록
        error_code = [["Sub_InPlayerDiff,", ",Can't find at log"], ["Sub_InPlayerDiff,", "Can't find at info"],
                      ["Sub_InTimeDiff,", ",incorrect"]]
        error_list = self.list_compare(sub_in_list, sub_log_in_list, error_list, error_code)
        error_code = [["Sub_OutPlayerDiff,", ",Can't find at log"], ["Sub_OutPlayerDiff,", "Can't find at info"],
                      ["Sub_OutTimeDiff,", ",incorrect"]]
        error_list = self.list_compare(sub_out_list, sub_log_out_list, error_list, error_code)

        # error file에 기록
        write_order = 'ErrorCode,ErrorValue,Detail\n'
        self.write_error(path_write_folder, 'data_difference.txt', write_order, error_list)

    def look_up_files_stable(self, path_raw_data, path_write, name_of_dir):
        # 기기로부터 안정적으로 정보를 받아오지 않았다면 error로 판단, 기록

        path_raw_data = common_os_helper.check_slash(path_raw_data)
        path_write = common_os_helper.check_slash(path_write)
        name_of_dir = common_os_helper.check_slash(name_of_dir)
        path_raw_data = path_raw_data + name_of_dir
        path_write = path_write + name_of_dir

        error_list = []
        for player in self.Players:
            file = open(path_raw_data+str(player[0])+'.csv', 'r')
            read_values = file.readlines()

            valid_count = 0.0
            in_time = player[1].split(':')
            out_time = player[2].split(':')
            total_count = float(out_time[0])*60+float(out_time[1])-float(in_time[0])*60-float(in_time[1])
            for i in range(1, len(read_values), 600):
                time_str = read_values[i].split(',')[3:5]
                time_str = time_str[0]+':'+time_str[1]
                if (self.time_comparision(time_str, player[1], 1) >= 0) and (
                        self.time_comparision(time_str, player[2], 1) <= 0):
                    valid_count += 1.0

            # player가 뛴 시간의 90% 이하로 정보가 있다면 불안정하다고 판단
            if valid_count/total_count <= 0.8:
                error_list.append("Unstable,"+str(player[0])+",Input value lost")
            elif valid_count/total_count <= 0.95:
                error_list.append("Unstable,"+str(player[0])+",Some input value lost")
            else:
                error_list.append("Stable,"+str(player[0]))

        # error file에 기록
        write_order = "ErrorCode,ErrorValue,Detail\n"
        self.write_error(path_write, "device_stability.txt", write_order, error_list)

    def look_up_player(self, path_read_xl, path_raw_data, path_write, name_of_dir):
        # excel file에 뛰었다고 기록된 선수가 log.csv에 뛰었다고 기록되어 있고, 데이터도 있는지 확인, 기록
        path_read_xl = common_os_helper.check_slash(path_read_xl)
        path_raw_data = common_os_helper.check_slash(path_raw_data)
        path_write = common_os_helper.check_slash(path_write)
        name_of_dir = common_os_helper.check_slash(name_of_dir)
        path_raw_data = path_raw_data+name_of_dir
        path_write = path_write+name_of_dir

        files_raw = glob.glob(path_raw_data+'*.csv')
        player_raw = []
        error_list = []
        for player in files_raw:
            player = player.replace('\\', '/')
            player = player.replace(path_raw_data, '')
            player = player.replace('.csv', '')
            # player의 등번호가 시리얼 번호와 맞지않아 번환되지 않았다면 NoMatch error
            if '_' in player or '(' in player:
                error_list.append("NoMatch,"+str(player)+",There is no player_num")
            else:
                player_raw.append([int(player), 0])

        # 경기가 home인지 away인지 모르기 때문에 둘 다 시도
        try:
            file_to_read_xl = openpyxl.load_workbook(path_read_xl+self.find_file_name(name_of_dir,'H'))
        except:
            file_to_read_xl = openpyxl.load_workbook(path_read_xl+self.find_file_name(name_of_dir, 'A'))
        active_sheet = file_to_read_xl.active
        player_info = active_sheet['F11':'F28']
        player_num = []
        for player in player_info:
            if player[0].value:
                player_num.append([int(player[0].value),0])

        # excel file에 기록된 선수들의 데이터가 있는지 raw data로 확인, 기록
        error_code = [["NoData,", ",Can't find data"], ["NoPlayer,", ",Can't find Player"]]
        error_list = self.list_compare(player_num, player_raw, error_list, error_code, False)

        # error file에 기록
        write_order = "ErrorCode,ErrorValue,Detail\n"
        self.write_error(path_write, 'player_difference.txt', write_order, error_list)

    def do_inspection(self, path_read_xl, path_read_log, path_raw_data, path_write, name_of_dir):
        # 3가지 주된 method를 실행하여 전체적인 inspection을 진행
        self.Players.clear()
        self.look_up_player(path_read_xl, path_raw_data, path_write, name_of_dir)
        self.find_diff_log(path_read_xl, path_read_log, path_write, name_of_dir)
        self.look_up_files_stable(path_raw_data, path_write, name_of_dir)
