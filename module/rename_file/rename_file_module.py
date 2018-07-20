import os
import glob
import time
import shutil
import openpyxl
from module.editdata.edit_data_module import EditData
from helper import common_os_helper

class Rename_file:

    def find_file_name(self, name_string, place, type = 'csv', state = '18'):
        # name_string 폴더의 이름을 string형태로 저장 ex) 'A-01_U18_서울'
        # place home이라면 H, away라면 A를 저장     ex)'H' 또는 'A'

        # serial 번호와 등번호가 적힌 .csv file의 이름을 찾음
        name_list = name_string.split('-')
        value_list = name_list[1].split('_')

        new_name = name_list[0] + '-' + str(int(value_list[0]))
        if value_list[1] == 'U17' or state == '17':
            new_name = new_name + '_U17'
        new_name = new_name+'_'+place+'_'+value_list[2]
        new_name = new_name.replace('/', '.'+type)
        return new_name

    def detect_games(self, path_read_csv, path_read_xl, name_of_dir):
        path_read_csv = common_os_helper.check_slash(path_read_csv)
        path_read_xl = common_os_helper.check_slash(path_read_xl)
        name_of_dir = common_os_helper.check_slash(name_of_dir)

        try:
            file_read_xl = openpyxl.load_workbook(path_read_xl+self.find_file_name(name_of_dir, 'H', 'xlsx'))
            file_read_xl17 = openpyxl.load_workbook(path_read_xl+self.find_file_name(name_of_dir, 'H', 'xlsx', '17'))
        except FileNotFoundError:
            try:
                file_read_xl = openpyxl.load_workbook(path_read_xl+self.find_file_name(name_of_dir, 'A', 'xlsx'))
                file_read_xl17 = openpyxl.load_workbook(path_read_xl+self.find_file_name(name_of_dir, 'A', 'xlsx', '17'))
            except FileNotFoundError:
                return

        active_sheet18 = file_read_xl.active
        active_sheet17 = file_read_xl17.active
        Start1_18 = active_sheet18['A7'].value
        End2_18 = active_sheet18['E7'].value
        Start1_17 = active_sheet17['A7'].value
        End2_17 = active_sheet17['E7'].value

        object_edit_data = EditData()
        object_edit_data.split_file(Start1_18, End2_18, Start1_17, End2_17,path_read_csv, name_of_dir)

    def rename_csv_file(self, path_target_folder, path_read_info_folder, path_read_xl, name_of_dir):
        # path_target_folder 이름을 변경하고자하는 target이 있는 folder path를 string으로 저장
        # path_read_info_folder serial_num,back_num으로 이루어진 .csv file이 있는 path를 string으로 저장
        # name_of_dir 파일을 읽고 쓸 dir name을 string으로 저장

        path_target_folder = common_os_helper.check_slash(path_target_folder)
        path_read_info_folder = common_os_helper.check_slash(path_read_info_folder)
        name_of_dir = common_os_helper.check_slash(name_of_dir)
        path_target_folder = path_target_folder+name_of_dir

        common_os_helper.create_dir(path_target_folder+'noneed/')

        # 현재 target인 팀이 home인지 away인지 모르기 때문에 두가지 경우 모두 시도
        try:
            file_read_info = open(path_read_info_folder+self.find_file_name(name_of_dir, 'H'), 'r')
        except FileNotFoundError:
            file_read_info = open(path_read_info_folder+self.find_file_name(name_of_dir, 'A'), 'r')

        info_values = file_read_info.readlines()
        file_read_info.close()
        info_values.sort()
        index = -1
        serial_num = -1
        last = -1
        count = 0

        files_target = glob.glob(path_target_folder+'*.csv')
        for file_target in files_target:
            try:
                file_target = file_target.replace('\\', '/')
                name_list = file_target.split('/')
                target_num = int(name_list[len(name_list)-1].split('_')[0])
            except:
                # filename이 변환되어 있는 경우 continue
                continue

            try:
                # target_num과 serial_num이 모두 정렬되어 있기 때문에 단순비교가 가능
                while target_num > serial_num:
                    index = index+1
                    serial_num = int(info_values[index].split(',')[0])
                    back_num = info_values[index].split(',')[1]
                    back_num = back_num[:len(back_num)-1]
                # 정렬된 상황에서 target_num < serial_num은 같은 번호를 가진 serial_num이 없다는 뜻이므로 noneed로 이동
                if target_num < serial_num:
                    shutil.move(file_target, path_target_folder+'noneed/'+name_list[len(name_list)-1])
                    continue
            except IndexError:
                shutil.move(file_target, path_target_folder+'noneed/'+name_list[len(name_list)-1])
                continue

            # 같은 serial number를 가진 파일이 여러개일 경우 (count)의 형식으로 이름에 추가
            if last == target_num:
                count = count+1
                re_num = back_num+'('+str(count)+')'
            else:
                count = 0
                re_num = back_num
                last = target_num

            name_list[len(name_list)-1] = re_num+'.csv'
            new_name = path_target_folder+name_list[len(name_list)-1]
            #os.rename(file_target, new_name)
            shutil.move(file_target, new_name)

        # self.detect_games(path_target_folder, path_read_xl, name_of_dir)